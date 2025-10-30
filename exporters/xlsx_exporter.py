#!/usr/bin/env python3
"""
Excel Exporter for CIA Factbook Refined Data

Exports refined country data to XLSX format with filtering and structured multi-value handling.
Supports filtering by countries and categories, and handles all structure types:
- simple: Single value fields
- key_value_pairs: Multiple key-value pairs
- key_sub_values: Keys with multiple sub-values
"""

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Any

# Add project root to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

# Module-level logger
logger = get_logger(__name__)


def load_filtered_data(
    snapshot_dir: str,
    country_filter: Optional[List[str]] = None,
    category_filter: Optional[List[str]] = None
) -> List[Dict[str, Any]]:
    """
    Load refined data with filters applied.

    Args:
        snapshot_dir: Path to snapshot directory
        country_filter: List of country slugs or None for all
        category_filter: List of categories or None for all

    Returns:
        List of filtered country data dicts
    """
    refined_dir = os.path.join(snapshot_dir, 'refined')

    if not os.path.exists(refined_dir):
        raise FileNotFoundError(f"Refined directory not found: {refined_dir}")

    # Get all refined JSON files
    json_files = [f for f in os.listdir(refined_dir) if f.endswith('.json')]

    if not json_files:
        raise ValueError(f"No refined JSON files found in {refined_dir}")

    logger.info(f"Found {len(json_files)} refined files")

    # Apply country filter
    if country_filter:
        filtered_files = [f for f in json_files if f.replace('.json', '') in country_filter]
        json_files = filtered_files
        logger.info(f"Filtered to {len(json_files)} countries: {country_filter}")

    # Load and filter data
    filtered_data = []

    for json_file in json_files:
        file_path = os.path.join(refined_dir, json_file)

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                country_data = json.load(f)

            # Apply category filter if specified
            if category_filter:
                filtered_fields = []
                for field in country_data['data']['fields']:
                    if field.get('category') in category_filter:
                        filtered_fields.append(field)

                # Update country data with filtered fields
                country_data['data']['fields'] = filtered_fields

                # Skip countries with no matching fields
                if not filtered_fields:
                    logger.debug(f"Skipping {json_file}: no fields match category filter")
                    continue

            filtered_data.append(country_data)

        except Exception as e:
            logger.error(f"Failed to load {json_file}: {e}")
            continue

    logger.info(f"Loaded {len(filtered_data)} countries after filtering")
    return filtered_data


def flatten_to_rows(country_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Convert refined structure to flat row format.

    Args:
        country_data: Single country refined data

    Returns:
        List of row dicts ready for Excel
    """
    rows = []

    country_slug = country_data['country_slug']
    country_name = country_data['data']['metadata']['name']

    fields = country_data['data']['fields']

    for field in fields:
        field_name = field['name']
        database_id = field['database_id']
        category = field.get('category', 'Unknown')
        structure_type = field['structure_type']
        has_ranking = field.get('has_ranking', False)

        if structure_type == 'simple':
            # Simple field: one row with single value
            for value_obj in field.get('values', []):
                row = {
                    'Country': country_name,
                    'Country Slug': country_slug,
                    'Field Name': field_name,
                    'Database ID': database_id,
                    'Category': category,
                    'Structure Type': structure_type,
                    'Key': '-',
                    'Value': value_obj.get('value', ''),
                    'Order': value_obj.get('order', 0),
                    'Has Ranking': has_ranking
                }
                rows.append(row)

        elif structure_type == 'key_value_pairs':
            # Key-value pairs: one row per pair
            for value_obj in field.get('values', []):
                row = {
                    'Country': country_name,
                    'Country Slug': country_slug,
                    'Field Name': field_name,
                    'Database ID': database_id,
                    'Category': category,
                    'Structure Type': structure_type,
                    'Key': value_obj.get('key', ''),
                    'Value': value_obj.get('value', ''),
                    'Order': value_obj.get('order', 0),
                    'Has Ranking': has_ranking
                }
                rows.append(row)

        elif structure_type == 'key_sub_values':
            # Key with sub-values: one row per sub-value
            for value_obj in field.get('values', []):
                key = value_obj.get('key', '')
                sub_values = value_obj.get('sub_values', [])

                for idx, sub_value in enumerate(sub_values):
                    row = {
                        'Country': country_name,
                        'Country Slug': country_slug,
                        'Field Name': field_name,
                        'Database ID': database_id,
                        'Category': category,
                        'Structure Type': structure_type,
                        'Key': key,
                        'Value': sub_value,
                        'Order': idx,
                        'Has Ranking': has_ranking
                    }
                    rows.append(row)

    logger.debug(f"Flattened {country_slug} to {len(rows)} rows")
    return rows


def format_excel(workbook: Workbook, sheet_name: str) -> None:
    """
    Apply formatting to Excel workbook.

    Args:
        workbook: openpyxl Workbook object
        sheet_name: Name of sheet to format
    """
    ws = workbook[sheet_name]

    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.alignment = header_alignment

    # Freeze first row
    ws.freeze_panes = 'A2'

    # Add filters to header row
    ws.auto_filter.ref = ws.dimensions

    # Alternate row colors (skip header)
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        if idx % 2 == 0:
            for cell in row:
                cell.fill = alt_row_fill

    # Auto-fit column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Set column width with min 10 and max 80
        adjusted_width = min(max(max_length + 2, 10), 80)
        ws.column_dimensions[column].width = adjusted_width

    logger.debug(f"Applied formatting to sheet '{sheet_name}'")


def create_excel_workbook(
    rows: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    output_path: str
) -> None:
    """
    Create XLSX workbook with data and summary.

    Args:
        rows: List of flattened row dicts
        metadata: Export metadata (countries, categories, etc.)
        output_path: Path to save XLSX file
    """
    workbook = Workbook()

    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # Create Fields sheet
    fields_sheet = workbook.create_sheet('Fields')

    # Define headers
    headers = [
        'Country',
        'Country Slug',
        'Field Name',
        'Database ID',
        'Category',
        'Structure Type',
        'Key',
        'Value',
        'Order',
        'Has Ranking'
    ]

    # Write headers
    fields_sheet.append(headers)

    # Write data rows
    for row in rows:
        fields_sheet.append([
            row['Country'],
            row['Country Slug'],
            row['Field Name'],
            row['Database ID'],
            row['Category'],
            row['Structure Type'],
            row['Key'],
            row['Value'],
            row['Order'],
            row['Has Ranking']
        ])

    logger.info(f"Created Fields sheet with {len(rows)} rows")

    # Create Summary sheet
    summary_sheet = workbook.create_sheet('Summary', 0)  # Insert as first sheet

    # Summary data
    summary_data = [
        ['Export Summary', ''],
        ['', ''],
        ['Export Date', datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')],
        ['Snapshot Date', metadata.get('snapshot_date', 'N/A')],
        ['', ''],
        ['Total Countries', metadata.get('total_countries', 0)],
        ['Total Fields', metadata.get('total_fields', 0)],
        ['Total Rows', len(rows)],
        ['', ''],
        ['Filters Applied', ''],
        ['Countries Filter', ', '.join(metadata.get('countries_filter', [])) or 'All countries'],
        ['Categories Filter', ', '.join(metadata.get('categories_filter', [])) or 'All categories'],
    ]

    # Write summary data
    for row in summary_data:
        summary_sheet.append(row)

    # Format summary sheet
    summary_sheet['A1'].font = Font(bold=True, size=14)
    for row in range(3, 13):
        summary_sheet[f'A{row}'].font = Font(bold=True)

    # Auto-fit columns
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 50

    logger.info("Created Summary sheet")

    # Apply formatting to Fields sheet
    format_excel(workbook, 'Fields')

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Save workbook
    workbook.save(output_path)
    logger.info(f"Saved Excel workbook to {output_path}")


def get_latest_snapshot() -> str:
    """
    Find most recent snapshot directory.

    Returns:
        Path to latest snapshot directory
    """
    snapshots_dir = "data/snapshots"

    try:
        if not os.path.exists(snapshots_dir):
            raise FileNotFoundError(f"Snapshots directory {snapshots_dir} not found")

        # Get all subdirectories
        snapshot_dirs = [
            d for d in os.listdir(snapshots_dir)
            if os.path.isdir(os.path.join(snapshots_dir, d))
        ]

        if not snapshot_dirs:
            raise ValueError("No snapshot directories found")

        # Sort by date (YYYY-MM-DD format should sort correctly)
        latest_snapshot = sorted(snapshot_dirs)[-1]
        latest_path = os.path.join(snapshots_dir, latest_snapshot)

        logger.info(f"Found latest snapshot: {latest_path}")
        return latest_path

    except Exception as e:
        logger.error(f"Failed to find latest snapshot: {e}")
        raise


def run(
    snapshot: Optional[str] = None,
    countries: Optional[List[str]] = None,
    categories: Optional[List[str]] = None,
    output: Optional[str] = None
) -> Dict[str, Any]:
    """
    Main execution function.

    Args:
        snapshot: Snapshot date or 'latest'
        countries: List of country slugs to export
        categories: List of categories to include
        output: Output file path

    Returns:
        Export summary statistics
    """
    start_time = datetime.now(timezone.utc)

    try:
        # Determine snapshot directory
        if snapshot:
            if snapshot == 'latest':
                snapshot_dir = get_latest_snapshot()
            else:
                snapshot_dir = os.path.join('data', 'snapshots', snapshot)
        else:
            snapshot_dir = get_latest_snapshot()

        if not os.path.exists(snapshot_dir):
            raise FileNotFoundError(f"Snapshot directory not found: {snapshot_dir}")

        # Extract snapshot date from path
        snapshot_date = os.path.basename(snapshot_dir)

        # Default output path
        if not output:
            output = f"exports/factbook_export_{snapshot_date}.xlsx"

        logger.info(f"Starting Excel export for snapshot: {snapshot_date}")

        print(f"\n{'='*60}")
        print(f"    CIA Factbook Excel Export")
        print(f"{'='*60}")
        print(f"Snapshot: {snapshot_date}")
        if countries:
            print(f"Countries: {', '.join(countries)}")
        else:
            print(f"Countries: All")
        if categories:
            print(f"Categories: {', '.join(categories)}")
        else:
            print(f"Categories: All")
        print(f"Output: {output}")
        print(f"{'='*60}\n")

        # Load filtered data
        print("Loading data...")
        country_data_list = load_filtered_data(snapshot_dir, countries, categories)

        if not country_data_list:
            raise ValueError("No data to export after applying filters")

        # Flatten all countries to rows
        print("Flattening data structure...")
        all_rows = []
        total_fields = 0

        for country_data in country_data_list:
            rows = flatten_to_rows(country_data)
            all_rows.extend(rows)
            total_fields += len(country_data['data']['fields'])

        logger.info(f"Flattened {len(country_data_list)} countries to {len(all_rows)} rows")

        # Prepare metadata
        metadata = {
            'snapshot_date': snapshot_date,
            'total_countries': len(country_data_list),
            'total_fields': total_fields,
            'countries_filter': countries or [],
            'categories_filter': categories or []
        }

        # Create Excel workbook
        print("Creating Excel workbook...")
        create_excel_workbook(all_rows, metadata, output)

        # Calculate duration
        duration = (datetime.now(timezone.utc) - start_time).total_seconds()

        # Print summary
        print(f"\n{'='*60}")
        print(f"    Export Complete")
        print(f"{'='*60}")
        print(f"Duration: {duration:.1f}s")
        print(f"Countries exported: {len(country_data_list)}")
        print(f"Fields exported: {total_fields}")
        print(f"Total rows: {len(all_rows)}")
        print(f"Output file: {output}")
        print(f"{'='*60}\n")

        # Return summary
        return {
            'success': True,
            'snapshot_date': snapshot_date,
            'countries_count': len(country_data_list),
            'fields_count': total_fields,
            'rows_count': len(all_rows),
            'output_file': output,
            'duration_seconds': duration
        }

    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        print(f"\nError: {e}")
        raise


def parse_arguments() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Export CIA Factbook refined data to Excel (XLSX)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export latest snapshot (all countries, all categories)
  python -m exporters.xlsx_exporter

  # Export specific snapshot
  python -m exporters.xlsx_exporter --snapshot 2025-10-27

  # Export specific countries
  python -m exporters.xlsx_exporter --countries france germany japan

  # Export specific categories (quote names with spaces)
  python -m exporters.xlsx_exporter --categories Economy Geography
  python -m exporters.xlsx_exporter --categories "Military and Security" "People and Society"

  # Combine filters and custom output
  python -m exporters.xlsx_exporter \\
    --snapshot 2025-10-27 \\
    --countries france germany \\
    --categories Economy Geography \\
    --output exports/europe_economy.xlsx

Available categories (quote names with spaces):
  Communications, Economy, Energy, Environment, Geography, Government,
  Introduction, "Military and Security", "People and Society", Space,
  Terrorism, "Transnational Issues", Transportation
        """
    )

    parser.add_argument(
        '--snapshot',
        type=str,
        help='Snapshot date (YYYY-MM-DD) or "latest" (default: latest)'
    )

    parser.add_argument(
        '--countries',
        nargs='+',
        help='List of country slugs to export (default: all)'
    )

    parser.add_argument(
        '--categories',
        nargs='+',
        help='List of categories to include - quote names with spaces (default: all)'
    )

    parser.add_argument(
        '--output',
        type=str,
        help='Output file path (default: exports/factbook_export_{date}.xlsx)'
    )

    return parser.parse_args()


def main():
    """Main entry point."""
    args = parse_arguments()

    try:
        result = run(
            snapshot=args.snapshot,
            countries=args.countries,
            categories=args.categories,
            output=args.output
        )

        sys.exit(0 if result['success'] else 1)

    except KeyboardInterrupt:
        print("\nExport interrupted by user.")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
